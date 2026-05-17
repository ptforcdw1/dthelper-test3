#!/usr/bin/env python3
"""
render_from_excel.py -- read an Excel workbook of template parameters and
write them into an ephemeral task's config.yaml and template.json.

Excel format (sheet 1):
  - Row 1: parameter names (one per column)
  - Row 2: parameter values (one per column, same order)

Flow:
  1. Read xlsx, build {param_name: value} from rows 1 and 2.
  2. Read tasks/<task-dir>/config.yaml; for each Excel param, overwrite
     configs[0].config.parameters[<name>].
  3. Read tasks/<task-dir>/template.json (Go-template-style JSON) and
     substitute every {{ .<param> }} placeholder with its Excel value
     (so the rendered file is plain JSON, easy to inspect in the logs).
  4. Write both files back in place.

Type coercion uses templates/<template>/request-schema.json:
  - "integer"  -> int(value)
  - "number"   -> float(value)
  - "boolean"  -> truthy parsing
  - default    -> str(value)
"""
import argparse
import json
import re
import sys

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required.  pip install pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def read_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    values  = [c.value for c in ws[2]] if ws.max_row >= 2 else []

    params = {}
    for h, v in zip(headers, values):
        if h is None or str(h).strip() == "":
            continue
        params[str(h).strip()] = v
    return params


def coerce(value, schema_type: str):
    if value is None or value == "":
        return None
    if schema_type == "integer":
        return int(value)
    if schema_type == "number":
        return float(value)
    if schema_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("true", "1", "yes", "y", "t")
    return str(value)


def render_config_yaml(path: str, params: dict, schema: dict) -> None:
    with open(path) as f:
        cfg = yaml.safe_load(f)

    target = cfg["configs"][0]["config"].setdefault("parameters", {})
    for name, value in params.items():
        stype = (schema.get("params", {}).get(name, {}) or {}).get("type")
        coerced = coerce(value, stype)
        if coerced is None:
            continue
        target[name] = coerced

    with open(path, "w") as f:
        yaml.safe_dump(cfg, f, default_flow_style=False, sort_keys=False)


def render_template_json(path: str, params: dict, schema: dict) -> None:
    with open(path) as f:
        text = f.read()

    coerced = {}
    for name, value in params.items():
        stype = (schema.get("params", {}).get(name, {}) or {}).get("type")
        c = coerce(value, stype)
        if c is not None:
            coerced[name] = c

    def repl(m: "re.Match") -> str:
        name = m.group(1)
        if name not in coerced:
            return m.group(0)
        v = coerced[name]
        if isinstance(v, bool):
            return "true" if v else "false"
        if isinstance(v, (int, float)):
            return str(v)
        # JSON-escape string contents but drop the wrapping quotes — template
        # placeholders are already inside quoted positions where needed.
        return json.dumps(str(v))[1:-1]

    new_text = re.sub(r"\{\{\s*\.([A-Za-z0-9_]+)\s*\}\}", repl, text)

    with open(path, "w") as f:
        f.write(new_text)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--template", required=True,
                    help="Template name under templates/")
    ap.add_argument("--xlsx", required=True,
                    help="Path to the user-uploaded params .xlsx")
    ap.add_argument("--task-dir", required=True,
                    help="Ephemeral task dir containing copies of config.yaml + template.json")
    args = ap.parse_args()

    params = read_xlsx(args.xlsx)
    print(f"Parameters read from {args.xlsx}:")
    for k, v in params.items():
        print(f"  {k} = {v!r}")

    schema_path = f"templates/{args.template}/request-schema.json"
    schema = {}
    try:
        with open(schema_path) as f:
            schema = json.load(f)
    except FileNotFoundError:
        print(f"  (no {schema_path} -- skipping type coercion)")

    render_config_yaml(f"{args.task_dir}/config.yaml", params, schema)
    render_template_json(f"{args.task_dir}/template.json", params, schema)
    print(f"Rendered {args.task_dir}/config.yaml and {args.task_dir}/template.json")


if __name__ == "__main__":
    main()
