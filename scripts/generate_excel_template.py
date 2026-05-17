#!/usr/bin/env python3
"""
generate_excel_template.py -- create an Excel workbook with one column per
template parameter. Row 1 = parameter names; row 2 = default values for the
user to overwrite.

Usage:
  python3 scripts/generate_excel_template.py \
      --template anomaly-service-responsetime-fix \
      --out      params-template.xlsx
"""
import argparse
import json
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--template", required=True,
                    help="Template subdirectory under templates/")
    ap.add_argument("--out", required=True,
                    help="Output .xlsx path")
    args = ap.parse_args()

    schema_path = f"templates/{args.template}/request-schema.json"
    try:
        with open(schema_path) as f:
            schema = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: {schema_path} not found.", file=sys.stderr)
        sys.exit(1)

    params = schema.get("params", {})
    if not params:
        print(f"ERROR: no 'params' in {schema_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parameters"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    for col, (name, spec) in enumerate(params.items(), start=1):
        h = ws.cell(row=1, column=col, value=name)
        h.font = header_font
        h.fill = header_fill

        default = spec.get("default", "")
        ws.cell(row=2, column=col, value=default)

        ws.column_dimensions[get_column_letter(col)].width = max(len(name) + 2, 16)

    ws.freeze_panes = "A2"

    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  Columns ({len(params)}): {', '.join(params.keys())}")
    print(f"  Row 2 is pre-filled with defaults — edit and re-upload.")


if __name__ == "__main__":
    main()
